from flask import Blueprint, request, jsonify
from sqlalchemy import func

from app.extensions import db
from app.models.product import Product
from app.notification_helper import create_notification

from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os

product_bp = Blueprint("product", __name__)


# ==========================================
# Add Product
# ==========================================
@product_bp.route("/products", methods=["POST"])
def add_product():
    data = request.get_json()

    name = data.get("name")
    category = data.get("category")
    price = data.get("price")
    stock = data.get("stock")
    status = data.get("status", "Active")
    description = data.get("description")

    if not name or not category or price is None or stock is None:
        return jsonify({
            "success": False,
            "message": "All required fields must be filled."
        }), 400

    product = Product(
        name=name,
        category=category,
        price=price,
        stock=stock,
        status=status,
        description=description,
    )

    db.session.add(product)
    db.session.commit()

# Create Notification
    create_notification(
        
        "New Product Added",
        f"{product.name} has been added successfully."
    )

    return jsonify({
    "success": True,
    "message": "Product added successfully.",
    "product": product.to_dict()
}), 201


# ==========================================
# Get Products (Pagination + Sorting)
# ==========================================
@product_bp.route("/products", methods=["GET"])
def get_products():

    page = request.args.get("page", default=1, type=int)
    per_page = request.args.get("per_page", default=10, type=int)
    sort = request.args.get("sort", "")

    query = Product.query

    if sort == "name_asc":
        query = query.order_by(Product.name.asc())

    elif sort == "name_desc":
        query = query.order_by(Product.name.desc())

    elif sort == "price_asc":
        query = query.order_by(Product.price.asc())

    elif sort == "price_desc":
        query = query.order_by(Product.price.desc())

    elif sort == "stock_asc":
        query = query.order_by(Product.stock.asc())

    elif sort == "stock_desc":
        query = query.order_by(Product.stock.desc())

    else:
        query = query.order_by(Product.id.desc())

    pagination = query.paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    return jsonify({
        "success": True,
        "products": [product.to_dict() for product in pagination.items],
        "page": pagination.page,
        "per_page": pagination.per_page,
        "total": pagination.total,
        "pages": pagination.pages,
        "has_next": pagination.has_next,
        "has_prev": pagination.has_prev
    }), 200


# ==========================================
# Update Product
# ==========================================
@product_bp.route("/products/<int:id>", methods=["PUT"])
def update_product(id):

    product = Product.query.get(id)

    if not product:
        return jsonify({
            "success": False,
            "message": "Product not found."
        }), 404

    data = request.get_json()

    product.name = data.get("name", product.name)
    product.category = data.get("category", product.category)
    product.price = data.get("price", product.price)
    product.stock = data.get("stock", product.stock)
    product.status = data.get("status", product.status)
    product.description = data.get("description", product.description)

    db.session.commit()

    return jsonify({
        "success": True,
        "message": "Product updated successfully.",
        "product": product.to_dict()
    }), 200


# ==========================================
# Delete Product
# ==========================================
@product_bp.route("/products/<int:id>", methods=["DELETE"])
def delete_product(id):

    product = Product.query.get(id)

    if not product:
        return jsonify({
            "success": False,
            "message": "Product not found."
        }), 404

    db.session.delete(product)
    db.session.commit()

    return jsonify({
        "success": True,
        "message": "Product deleted successfully."
    }), 200

# ==========================================
# Bulk Delete Products
# ==========================================
@product_bp.route("/products/bulk-delete", methods=["DELETE"])
def bulk_delete_products():

    data = request.get_json()

    ids = data.get("ids", [])

    if not ids:
        return jsonify({
            "success": False,
            "message": "No products selected."
        }), 400

    Product.query.filter(
        Product.id.in_(ids)
    ).delete(synchronize_session=False)

    db.session.commit()

    return jsonify({
        "success": True,
        "message": f"{len(ids)} products deleted successfully."
    }), 200


# ==========================================
# Import Products from Excel
# ==========================================
@product_bp.route("/products/import", methods=["POST"])
def import_products():

    if "file" not in request.files:
        return jsonify({
            "success": False,
            "message": "No file uploaded."
        }), 400

    file = request.files["file"]

    if file.filename == "":
        return jsonify({
            "success": False,
            "message": "Please select an Excel file."
        }), 400

    filename = secure_filename(file.filename)

    upload_folder = "uploads"
    os.makedirs(upload_folder, exist_ok=True)

    filepath = os.path.join(upload_folder, filename)
    file.save(filepath)

    workbook = load_workbook(filepath)
    sheet = workbook.active

    imported = 0

    # Skip Header Row
    for row in sheet.iter_rows(min_row=2, values_only=True):

        if not row[0]:
            continue

        product = Product(
            name=row[0],
            category=row[1],
            price=row[2],
            stock=row[3],
            status=row[4] if row[4] else "Active",
            description=row[5]
        )

        db.session.add(product)
        imported += 1

    db.session.commit()

    os.remove(filepath)

    return jsonify({
        "success": True,
        "message": f"{imported} products imported successfully.",
        "imported": imported
    }), 201


# ==========================================
# Product Summary API
# ==========================================
@product_bp.route("/products/summary", methods=["GET"])
def product_summary():

    total_products = Product.query.count()

    categories = db.session.query(
        func.count(func.distinct(Product.category))
    ).scalar()

    in_stock = Product.query.filter(Product.stock > 0).count()

    out_of_stock = Product.query.filter(Product.stock <= 0).count()

    return jsonify({
        "success": True,
        "summary": {
            "total_products": total_products,
            "categories": categories,
            "in_stock": in_stock,
            "out_of_stock": out_of_stock
        }
    }), 200